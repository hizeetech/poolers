from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.utils import timezone
from django.http import JsonResponse
from django.http import HttpResponse
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from betting.models import ActivityLog, BettingPeriod, User
from .models import FraudAlert, AlertAffectedUser, InvestigationCase, AdminActionLog
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from .services import DashboardService
from .forecasting import ForecastingService
from .reporting import ReportingService

def is_admin_or_executive(user):
    return user.is_authenticated and (user.is_superuser or user.user_type in ['admin', 'account_user'])

@login_required
@user_passes_test(is_admin_or_executive)
def uip_dashboard(request):
    timeframe = request.GET.get('timeframe', 'daily')
    if timeframe not in ['daily', 'weekly', 'monthly']:
        timeframe = 'daily'

    start_date = request.GET.get("start_date") or None
    end_date = request.GET.get("end_date") or None
    period_id = request.GET.get("period_id") or None

    now = timezone.localtime(timezone.now())
    selected_period = None
    if period_id:
        try:
            selected_period = BettingPeriod.objects.filter(id=int(period_id)).first()
        except Exception:
            selected_period = None

    if selected_period:
        sd = selected_period.start_date
        ed = selected_period.end_date
    elif start_date and end_date:
        try:
            sd = timezone.datetime.fromisoformat(start_date).date()
            ed = timezone.datetime.fromisoformat(end_date).date()
        except Exception:
            sd = now.date()
            ed = now.date()
    else:
        if timeframe == "weekly":
            sd = (now - timedelta(days=now.weekday())).date()
            ed = now.date()
        elif timeframe == "monthly":
            sd = now.replace(day=1).date()
            ed = now.date()
        else:
            sd = now.date()
            ed = now.date()

    start_time = timezone.make_aware(timezone.datetime.combine(sd, timezone.datetime.min.time()))
    end_time = timezone.make_aware(timezone.datetime.combine(ed, timezone.datetime.max.time()))

    metrics = DashboardService.get_live_metrics(timeframe=timeframe)
    leaderboards = DashboardService.get_agent_leaderboards(start_time, end_time, limit=10)
    leaderboard = leaderboards.get("top_turnover") or []
    recent_activity = DashboardService.get_recent_activity()
    
    # Fraud Alerts Filtering
    fraud_alerts = FraudAlert.objects.prefetch_related('affected_user_details__user')
    
    alert_type = request.GET.get('alert_type')
    if alert_type:
        fraud_alerts = fraud_alerts.filter(alert_type=alert_type)
        
    severity = request.GET.get('severity')
    if severity:
        fraud_alerts = fraud_alerts.filter(severity=severity)
        
    status = request.GET.get('status')
    if status:
        fraud_alerts = fraud_alerts.filter(status=status)
        
    search = request.GET.get('search')
    if search:
        fraud_alerts = fraud_alerts.filter(
            Q(description__icontains=search) |
            Q(affected_users__email__icontains=search) |
            Q(affected_users__username__icontains=search) |
            Q(related_ips__contains=[search])
        ).distinct()
    
    fraud_alerts = fraud_alerts.order_by('-timestamp')[:50]
    
    betting_periods = BettingPeriod.objects.filter(is_active=True).order_by('-start_date')[:10]
    context = {
        'metrics': metrics,
        'leaderboard': leaderboard,
        'recent_activity': recent_activity,
        'fraud_alerts': fraud_alerts,
        'betting_periods': betting_periods,
        'page_title': 'Unified Intelligence Platform',
        'current_timeframe': timeframe,
        'leaderboard_start_date': sd.isoformat(),
        'leaderboard_end_date': ed.isoformat(),
        'leaderboard_period_id': str(selected_period.id) if selected_period else "",
    }
    return render(request, 'uip/dashboard.html', context)

@login_required
@user_passes_test(is_admin_or_executive)
def uip_financials(request):
    metrics = DashboardService.get_financial_metrics()
    context = {
        'metrics': metrics,
        'page_title': 'Financial Intelligence'
    }
    return render(request, 'uip/financials.html', context)

@login_required
@user_passes_test(is_admin_or_executive)
def uip_analytics(request):
    metrics = DashboardService.get_analytics_metrics()
    context = {
        'metrics': metrics,
        'page_title': 'Agent & User Analytics'
    }
    return render(request, 'uip/analytics.html', context)

@login_required
@user_passes_test(is_admin_or_executive)
def uip_risk(request):
    metrics = DashboardService.get_risk_metrics()
    context = {
        'metrics': metrics,
        'page_title': 'Risk & Fraud Monitoring'
    }
    return render(request, 'uip/risk.html', context)

@login_required
@user_passes_test(is_admin_or_executive)
def uip_forecasting(request):
    turnover_forecast = ForecastingService.predict_turnover()
    peak_periods = ForecastingService.identify_peak_periods()
    
    context = {
        'forecast': turnover_forecast,
        'peak_periods': peak_periods,
        'page_title': 'Forecasting & Projections'
    }
    return render(request, 'uip/forecasting.html', context)

@login_required
@user_passes_test(is_admin_or_executive)
def uip_reports(request):
    return render(request, 'uip/reports.html', {'page_title': 'Automated Reporting'})

@login_required
@user_passes_test(is_admin_or_executive)
def uip_audit(request):
    # Filtering
    logs = ActivityLog.objects.all().order_by('-timestamp')
    
    email = request.GET.get('email')
    if email:
        logs = logs.filter(user__email__icontains=email)
        
    action_type = request.GET.get('action_type')
    if action_type:
        logs = logs.filter(action_type=action_type)
        
    date_str = request.GET.get('date')
    if date_str:
        logs = logs.filter(timestamp__date=date_str)

    paginator = Paginator(logs, 50) # Show 50 contacts per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'logs': page_obj,
        'page_title': 'Audit & Compliance Logs'
    }
    return render(request, 'uip/audit.html', context)

@login_required
@user_passes_test(is_admin_or_executive)
def serial_number_analytics(request):
    """
    API endpoint for real-time serial number frequency chart.
    """
    start_date = request.GET.get('start_date') or None
    end_date = request.GET.get('end_date') or None
    scope = request.GET.get('scope') or 'all'
    user_id = request.GET.get('user_id') or None
    period_id = request.GET.get('period_id') or None

    data = DashboardService.get_serial_number_frequency(
        start_date=start_date,
        end_date=end_date,
        scope=scope,
        user_id=user_id,
        period_id=period_id
    )
    return JsonResponse(data)

@login_required
@user_passes_test(is_admin_or_executive)
def export_financials(request):
    start_date = request.GET.get('start_date') or timezone.now().date()
    end_date = request.GET.get('end_date') or timezone.now().date()
    return ReportingService.export_financial_report(start_date, end_date)

@login_required
@user_passes_test(is_admin_or_executive)
def export_agents(request):
    start_date = request.GET.get('start_date') or timezone.now().date()
    end_date = request.GET.get('end_date') or timezone.now().date()
    return ReportingService.export_agent_performance_report(start_date, end_date)

@login_required
@user_passes_test(is_admin_or_executive)
def export_audit(request):
    start_date = request.GET.get('start_date') or timezone.now().date()
    end_date = request.GET.get('end_date') or timezone.now().date()
    return ReportingService.export_audit_log_report(start_date, end_date)


@login_required
@user_passes_test(is_admin_or_executive)
def uip_leaderboards(request):
    timeframe = request.GET.get("timeframe") or "daily"
    start_date = request.GET.get("start_date") or None
    end_date = request.GET.get("end_date") or None

    now = timezone.localtime(timezone.now())
    if start_date and end_date:
        try:
            sd = timezone.datetime.fromisoformat(start_date).date()
            ed = timezone.datetime.fromisoformat(end_date).date()
        except Exception:
            sd = now.date()
            ed = now.date()
    else:
        if timeframe == "weekly":
            sd = (now - timedelta(days=now.weekday())).date()
            ed = now.date()
        elif timeframe == "monthly":
            sd = now.replace(day=1).date()
            ed = now.date()
        elif timeframe == "yearly":
            sd = now.replace(month=1, day=1).date()
            ed = now.date()
        else:
            sd = now.date()
            ed = now.date()

    start_time = timezone.make_aware(timezone.datetime.combine(sd, timezone.datetime.min.time()))
    end_time = timezone.make_aware(timezone.datetime.combine(ed, timezone.datetime.max.time()))

    data = DashboardService.get_agent_leaderboards(start_time, end_time, limit=50)

    if (request.GET.get("export") or "").lower() == "xlsx":
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Deposits"
        ws2 = wb.create_sheet("Turnover")
        ws3 = wb.create_sheet("ProfitMargin")

        def write_sheet(ws, header, rows):
            ws.append(header)
            for r in rows:
                ws.append(r)
            for i, col in enumerate(header, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(col)) + 2)

        write_sheet(
            ws1,
            ["Rank", "Agent Email", "Total Deposits", "Turnover", "Tickets"],
            [
                [idx + 1, a.email, float(getattr(a, "total_deposits", 0) or 0), float(getattr(a, "total_turnover", 0) or 0), int(getattr(a, "tickets_sold", 0) or 0)]
                for idx, a in enumerate(data.get("top_deposits") or [])
            ],
        )
        write_sheet(
            ws2,
            ["Rank", "Agent Email", "Turnover", "Tickets", "Winnings Paid"],
            [
                [idx + 1, a.email, float(getattr(a, "total_turnover", 0) or 0), int(getattr(a, "tickets_sold", 0) or 0), float(getattr(a, "winnings_paid", 0) or 0)]
                for idx, a in enumerate(data.get("top_turnover") or [])
            ],
        )
        write_sheet(
            ws3,
            ["Rank", "Agent Email", "Profit Margin %", "Revenue", "Turnover", "Winnings", "Tickets"],
            [
                [idx + 1, row["agent"].email, round(float(row["margin"]), 2), round(float(row["revenue"]), 2), round(float(row["turnover"]), 2), round(float(row["winnings"]), 2), int(row["tickets"])]
                for idx, row in enumerate(data.get("top_margin") or [])
            ],
        )

        resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="agent_leaderboards_{sd}_{ed}.xlsx"'
        wb.save(resp)
        return resp

    context = {
        "page_title": "Leaderboards",
        "timeframe": timeframe,
        "start_date": sd,
        "end_date": ed,
        "leaderboards": data,
    }
    return render(request, "uip/leaderboards.html", context)

@login_required
@user_passes_test(is_admin_or_executive)
def investigation_user_action(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        action = data.get('action')
        user_id = data.get('user_id')
        alert_id = data.get('alert_id')
        
        try:
            target_user = User.objects.get(id=user_id)
            alert = FraudAlert.objects.get(id=alert_id)
            
            if action == 'freeze':
                target_user.is_active = False
                target_user.save()
                AdminActionLog.objects.create(
                    alert=alert, admin=request.user, 
                    action=f"Froze user {target_user.email}",
                    notes=data.get('notes', '')
                )
            elif action == 'unfreeze':
                target_user.is_active = True
                target_user.save()
                AdminActionLog.objects.create(
                    alert=alert, admin=request.user, 
                    action=f"Unfroze user {target_user.email}",
                    notes=data.get('notes', '')
                )
            elif action == 'suspend_betting':
                target_user.is_locked = True
                target_user.lock_reason = "Suspended due to fraud investigation"
                target_user.save()
                AdminActionLog.objects.create(
                    alert=alert, admin=request.user, 
                    action=f"Suspended betting for {target_user.email}",
                    notes=data.get('notes', '')
                )
            elif action == 'unsuspend_betting':
                target_user.is_locked = False
                target_user.lock_reason = ""
                target_user.save()
                AdminActionLog.objects.create(
                    alert=alert, admin=request.user, 
                    action=f"Unsuspended betting for {target_user.email}",
                    notes=data.get('notes', '')
                )
            
            return JsonResponse({'status': 'success', 'message': f"Action {action} applied successfully."})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=405)

@login_required
@user_passes_test(is_admin_or_executive)
def update_fraud_alert_status(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        alert_id = data.get('alert_id')
        new_status = data.get('status')
        
        try:
            alert = FraudAlert.objects.get(id=alert_id)
            old_status = alert.status
            alert.status = new_status
            alert.save()
            
            AdminActionLog.objects.create(
                alert=alert, admin=request.user, 
                action=f"Changed status from {old_status} to {new_status}",
                notes=data.get('notes', '')
            )
            
            return JsonResponse({'status': 'success', 'message': "Alert status updated."})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=405)

@login_required
@user_passes_test(is_admin_or_executive)
def add_fraud_alert_note(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        alert_id = data.get('alert_id')
        note = data.get('note')
        
        try:
            alert = FraudAlert.objects.get(id=alert_id)
            AdminActionLog.objects.create(
                alert=alert, admin=request.user, 
                action="Added investigation note",
                notes=note
            )
            return JsonResponse({'status': 'success', 'message': "Note added."})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=405)

@login_required
@user_passes_test(is_admin_or_executive)
def export_investigation_report(request, alert_id):
    try:
        alert = FraudAlert.objects.get(id=alert_id)
        details = alert.affected_user_details.select_related('user').all()
        
        format = request.GET.get('format', 'xlsx')
        
        if format == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = "Investigation Report"
            
            ws.append(["Alert ID", f"#{alert.id}"])
            ws.append(["Alert Type", alert.get_alert_type_display()])
            ws.append(["Severity", alert.severity.upper()])
            ws.append(["Status", alert.get_status_display()])
            ws.append(["Timestamp", alert.timestamp.strftime("%Y-%m-%d %H:%M")])
            ws.append(["Description", alert.description])
            ws.append([])
            
            headers = ["Username", "Email", "IP Address", "Risk Score", "Wallet Balance", "Total Deposits", "Total Withdrawals", "Total Bets"]
            ws.append(headers)
            
            for d in details:
                ws.append([
                    d.user.username or d.user.get_full_name(),
                    d.user.email,
                    d.ip_address,
                    d.risk_score,
                    float(d.wallet_balance),
                    float(d.total_deposits),
                    float(d.total_withdrawals),
                    d.total_bets_count
                ])
                
            resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = f'attachment; filename="investigation_report_{alert_id}.xlsx"'
            wb.save(resp)
            return resp
        
        elif format == 'csv':
            import csv
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="investigation_report_{alert_id}.csv"'
            writer = csv.writer(resp)
            writer.writerow(["Alert Type", alert.get_alert_type_display()])
            writer.writerow(["Severity", alert.severity.upper()])
            writer.writerow(["Description", alert.description])
            writer.writerow([])
            writer.writerow(["Username", "Email", "IP Address", "Risk Score", "Wallet Balance", "Total Bets"])
            for d in details:
                writer.writerow([
                    d.user.username or d.user.get_full_name(),
                    d.user.email,
                    d.ip_address,
                    d.risk_score,
                    d.wallet_balance,
                    d.total_bets_count
                ])
            return resp

    except FraudAlert.DoesNotExist:
        return HttpResponse("Alert not found", status=404)
    except Exception as e:
        return HttpResponse(str(e), status=500)

@login_required
@user_passes_test(is_admin_or_executive)
def sync_investigation_alerts(request):
    if request.method == 'POST':
        try:
            from django.core.management import call_command
            call_command('spool_fraud_alerts')
            return JsonResponse({'status': 'success', 'message': 'Alerts spooled successfully.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=405)
